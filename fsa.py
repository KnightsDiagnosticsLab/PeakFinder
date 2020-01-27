import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from string import ascii_uppercase
import random
import re
import win32com.client as win32
import easygui

def convert_xls_to_xlsx(file_path=None):
	if file_path is None:
		file_path = easygui.fileopenbox(msg='Select .xls file to convert to .xlsx')
	file_path = file_path.replace('/', '\\')
	# print(file_path)
	excel = win32.gencache.EnsureDispatch('Excel.Application')
	wb = excel.Workbooks.Open(file_path)
	# xl = win32.Dispatch("Excel.Application")
	# print(xl.ActiveWorkbook.FullName)
	# print(wb.Name)
	wb.Name = str(wb.Name) + 'x'
	# print(wb.Name, '\n\n\n\n\n\n')
	new_file_path = file_path + 'x'
	wb.SaveAs(Filename=new_file_path, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
	wb.Close()                               #FileFormat = 56 is for .xls extension
	excel.Application.Quit()
	return new_file_path

def get_col_to_drop(df):
	col_to_drop = []
	for col in df.columns:
		# if 'Unnamed' in str(col) and df[col].isnull().all():
		if df[col].isnull().all():
			col_to_drop.append(col)
	return col_to_drop


def df_to_wb(df):
	wb = openpyxl.Workbook()
	ws = wb.active

	# print(df)
	for row_num, row_name in enumerate(df.index.tolist()):
		for col_num, col_name in enumerate(df.columns.tolist()):
			ws.cell(row=row_num + 1, column=col_num + 1).value = df.iloc[row_num, col_num]

	return wb


def use_csv_module(filename):
	with open(filename, newline='') as csvfile:
		dialect = csv.Sniffer().sniff(csvfile.read(1024))
		csvfile.seek(0)
		reader = csv.reader(csvfile, dialect)
		l = [r for r in reader]
	headers = l.pop(0)
	df = pd.DataFrame(l, columns=headers)
	df.replace(r'^\s*$', pd.np.nan, regex=True, inplace=True)
	# print(df)
	return df


def fix_formatting(filename, header=None, case_name=None):
	''' Helper function '''
	def border_add(border, top=None, right=None, left=None, bottom=None):
		if top is None:
			top = border.top
		if left is None:
			left = border.left
		if right is None:
			right = border.right
		if bottom is None:
			bottom = border.bottom
		return openpyxl.styles.Border(
			top=top, left=left, right=right, bottom=bottom)

	# print(filename)
	assert filename.endswith('.xlsx')
	thin = Side(border_style='thin')
	medium = Side(border_style='medium')
	df = pd.read_excel(filename)
	wb = openpyxl.load_workbook(filename)
	ws = wb.worksheets[0]

	'''	Insert header '''
	if header is not None:
		ws.oddHeader = header

	''' By default make all cells horizontal='center', except for the first and last two columns
	'''
	cells = [ws[c + str(r)] for c in ascii_uppercase[1:ws.max_column-1]
			 for r in range(1, ws.max_row + 1)]
	for cell in cells:
		cell.alignment = Alignment(horizontal='center')

	''' Make first column bold and right aligned '''
	cells = [ws['A' + str(r)] for r in range(3, ws.max_row - 1)]
	for cell in cells:
		# cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal='right')
		cell.border = border_add(cell.border, right=medium)

	''' Make first two rows bold and left aligned '''
	cells = [ws[c + str(i)] for i in range(1, 3)
			 for c in ascii_uppercase[0:ws.max_column]]
	for cell in cells:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal='left')

	'''	Get locations of 'Allele' '''
	allele_ij = []
	for i in df.index:
		for j, v in enumerate(df.iloc[i]):
			if v == 'Allele':
				allele_ij.append([i, j])

	''' Apply medium thickness based on which cells have the word 'Allele'
	'''
	for i, j in allele_ij:
		for k in range(0, ws.max_column):
			cell = ws[ascii_uppercase[k] + str(i + 3)]
			cell.border = border_add(cell.border, bottom=medium)

		for k in range(0, j, 2):
			cell = ws[ascii_uppercase[k] + str(i + 2)]
			cell.border = border_add(cell.border, right=medium)

			cell = ws[ascii_uppercase[k] + str(i + 3)]
			cell.border = border_add(cell.border, right=medium)

			cell = ws[ascii_uppercase[k + j] + str(i + 2)]
			cell.border = border_add(cell.border, right=medium)

			cell = ws[ascii_uppercase[k + j] + str(i + 3)]
			cell.border = border_add(cell.border, right=medium)

		cell = ws[ascii_uppercase[2 * j] + str(i + 2)]
		cell.border = border_add(cell.border, right=medium)

		cell = ws[ascii_uppercase[2 * j] + str(i + 3)]
		cell.border = border_add(cell.border, right=medium)

	''' Add in case_name '''
	if case_name is not None:
		loc = location_of_value(ws, 'Post-T:')
		if loc is not None:
			cell = ws[chr(ord(loc[0]) + 1) + str(loc[1])]
			cell.value = case_name

	if header is not None:
		ws.oddHeader = header
	ws.sheet_view.view = 'pageLayout'
	openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
		ws, paper_size=1, orientation='landscape')
	ws.page_margins.bottom = 0.5
	ws.page_margins.top = 0.5
	ws.page_margins.left = 0.5
	ws.page_margins.right = 0.5
	ws.page_margins.header = 0.1

	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.page_setup.fitToHeight = False

	ws.page_setup.fitToWidth = 1

	wb.save(filename)
	wb.close()  # FileFormat = 56 is for .xls extension

def replace_cell_values(ws, replacement_dict, regex=False):
	for r in range(1, ws.max_row + 1):
		for c in range(1, ws.max_column + 1):
			cell = ws.cell(row=r, column=c)
			for old, new in replacement_dict.items():
				if regex:
					if re.fullmatch(old, str(cell.value), flags=re.IGNORECASE) is not None:
						cell.value = new
				elif cell.value == old:
					cell.value = new
	return ws


def locations_of_value(ws, val):
	locs = []
	for r in range(1, ws.max_row + 1):
		for c in range(1, ws.max_column + 1):
			cell = ws.cell(row=r, column=c)
			if val == cell.value:
				locs.append(cell.coordinate)
	return locs


def idx_of_value(ws, val):
	idx = [None, None]
	for r in range(1, ws.max_row + 1):
		for c in range(1, ws.max_column + 1):
			cell = ws.cell(row=r, column=c)
			if val == cell.value:
				idx = [cell.column, cell.row]
				return idx
	return idx


def cell_with_value(ws, val):
	for r in range(1, ws.max_row + 1):
		for c in range(1, ws.max_column + 1):
			cell = ws.cell(row=r, column=c)
			if cell.value == val:
				return cell
	return None


def location_of_value(ws, val):
	loc = None
	# print('\t************ ws.max_column = {}'.format(ws.max_column))
	# print('\t************ ws.max_row = {}'.format(ws.max_row))
	for r in range(1, ws.max_row + 1):
		for c in range(1, ws.max_column + 1):
			# c = ascii_uppercase[i]
			# cell = ws[c + str(j)]
			cell = ws.cell(row=r, column=c)
			# print('c = {}'.format(c))
			# print('c+1 = {}'.format(c))
			# print('cell.coordinate = {}'.format(cell.coordinate))
			if val == cell.value:
				loc = cell.coordinate
				# print('\t\t******** {}'.format(loc))
				# loc = (c, j)
				# print('\t\t******** {}'.format(loc))
				# loc = (i, j)
				# print('loc of {} = {}'.format(val, loc))
				return loc
	return loc

def replace_cell_ref_with_value(df):
	# print('now inside replace_cell_ref_with_value')
	wb = df_to_wb(df)
	ws = wb.worksheets[0]
	# allele_locs = locations_of_value(ws, 'Allele')
	# host_loc = location_of_value(ws, 'ENG Host:')
	# donor_loc = location_of_value(ws, 'DEG Donor:')
	for r in range(1,ws.max_row+1):
		for c in range(1,ws.max_column+1):
			cell = ws.cell(row=r, column=c)
			if re.fullmatch('=[A-Z]+\d+',str(cell.value)):
				coor = str(cell.value).replace('=','')
				cell.value = ws[coor].value
	df = pd.DataFrame(ws.values)
	# print(df)
	return df


def build_results_dict(df=None):
	peaks = {}

	# if not isinstance(df, pd.DataFrame):
	# 	filename = easygui.fileopenbox(
	# 		msg='Select results file')
	# 	if filename is None:
	# 		exit()
	# 	df = use_csv_module(filename)

	if isinstance(df, pd.DataFrame):
		df = df[['Sample File Name', 'Marker', 'Allele', 'Area']]

		'''	Get rid of peaks that aren't assigned an allele '''
		df = df.dropna(axis=0, how='any', inplace=False)

		'''	Get rid of OL (off ladder) peaks '''
		df = df[df['Allele'] != 'OL']
		df = df.reset_index(drop=True, inplace=False)

		fnames = set()
		for i in df.index:
			file_name = str(df.iloc[i]['Sample File Name'])
			fnames.add(file_name)
			locus = str(df.iloc[i]['Marker'])
			allele = str(df.iloc[i]['Allele'])
			key = (file_name, locus, allele)
			peaks[key] = peaks.get(key, 0) + int(df.iloc[i]['Area'])
	# for k,v in peaks.items():
	# 	print(k,v)
	return peaks
