#!/usr/bin/env python3

import re
import os
import sys
import pandas as pd
import easygui
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter, get_column_interval
import win32com.client as win32
import csv
from fsa import *
import random

from pprint import pprint

pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 1000)
pd.set_option('display.max_rows', 50)


''' Helper function '''
def string_to_number(s):
	try:
		s = float(s)
	except:
		pass
	return s


''' Helper function '''
def format_value(cell_value):
	if cell_value is not None:
		cell_value_formatted = str(cell_value).replace('.0', '')
	else:
		cell_value_formatted = None
	return cell_value_formatted


def make_template_wb(file_path=None, host_case=None, donor_cases=[], df_cases={}, host_name=''):

	markers = ['D8S1179', 'D21S11', 'D7S820', 'CSF1PO', 'D3S1358', 'TH01', 'D13S317', 'D16S539', 'D2S1338', 'D19S433', 'vWA', 'TPOX', 'D18S51', 'AMEL', 'D5S818', 'FGA']
	# host_case = host_case
	df_host = df_cases.get(host_case, pd.DataFrame())
	df_host = df_host.loc[df_host['Selected'] == True]
	# print('df_host')
	# print(df_host)

	# donor_cases = donor_cases.value
	donors = {}
	for donor_case in donor_cases:
		df_donor = df_cases[donor_case]
		df_donor = df_donor.loc[df_donor['Selected'] == True]
		donors[donor_case] = df_donor.copy(deep=True)

	wb = openpyxl.Workbook()
	ws = wb.active

	'''	Header, etc. '''
	ws.oddHeader.center.text = host_name
	# ws.oddHeader.center.size = 14
	ws.cell(row=2, column=1, value='Marker')
	ws.cell(row=2, column=2, value='Host')
	host_case_abbrev = re.sub(r'_PTE.*$', '', host_case)
	ws.cell(row=1, column=2, value=host_case_abbrev)
	for i, donor_case in enumerate(donor_cases):
		donor_case_abbrev = re.sub(r'_PTE.*$', '', donor_case)
		c = 4 + 2*i
		ws.cell(row=1, column=c, value=donor_case_abbrev)
		if len(donor_cases) == 1:
			donor_num = 'Donor'
		else:
			donor_num = 'Donor ' + str(i+1)
		ws.cell(row=2, column=c, value=donor_num)

	'''	Markers & Alleles '''
	for i, marker in enumerate(markers):
		r = 1+(i+1)*2
		df_marker = df_host.loc[df_host['Marker'] == marker]
		ws.cell(row=r, column=1, value=marker)
		# print(df_marker)
		for j, allele in enumerate(df_marker['Allele']):
			# print(marker, j, allele)
			allele = string_to_number(allele)
			c = j+2
			ws.cell(row=r, column=c, value=allele)
		
		for j, donor in enumerate(donors.keys()):
			df_donor = donors[donor]
			df_marker = df_donor.loc[df_donor['Marker'] == marker]
			for k, allele in enumerate(df_marker['Allele']):
				allele = string_to_number(allele)
				c = (4 + 2*j) + k
				ws.cell(row=r, column=c, value=allele)

	''' Add in column of Allele/Area '''
	caa = 2*len(donors.keys()) + 4
	for i, marker in enumerate(markers):
		r1 = 1+(i+1)*2
		r2 = r1 + 1
		ws.cell(row=r1, column=caa, value='Allele')
		ws.cell(row=r2, column=caa, value='Area')

	''' Copy columns '''
	for i in range(2,caa,2):
		for r in range(2,ws.max_row + 1):
			c1 = 2*caa - (i+1)
			c2 = c1 + 1
			ws.cell(row=r, column=c1, value=ws.cell(row=r, column=i).value)
			ws.cell(row=r, column=c2, value=ws.cell(row=r, column=i+1).value)

	'''	Add in % Host & Forumula columns '''
	ws.cell(row=2, column=2*caa - 1, value='% Host')
	ws.cell(row=2, column=2*caa, value='Formula')
	ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Host')
	ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Donor')

	wb = insert_equations(wb=wb, file_path=file_path, host_case=host_case, donor_cases=donor_cases, df_cases=df_cases, host_name=host_name)

	if file_path is not None:
		if file_path.endswith('.xlsx'):
			'''	Save the file before running fix_formatting '''
			wb.save(file_path)
			wb.close()

			'''	Fix formatting '''
			fix_formatting(file_path)
			print('Done saving {}'.format(file_path))

	wb.close()

	return wb

def insert_equations(wb=None, file_path=None, host_case=None, donor_cases=[], df_cases={}, host_name=''):
	'''
	*************************************************
	THIS IS WHERE WE NEED TO MAKE EDITS FOR 2+ DONORS
	*************************************************
	'''
	markers = ['D8S1179', 'D21S11', 'D7S820', 'CSF1PO', 'D3S1358', 'TH01', 'D13S317', 'D16S539', 'D2S1338', 'D19S433', 'vWA', 'TPOX', 'D18S51', 'AMEL', 'D5S818', 'FGA']

	ws = wb.active

	allele_cell = first_cell_with_value(ws, 'Allele')

	caa = allele_cell.column
	raa = allele_cell.row

	if len(donor_cases) == 1:

		'''	Add the actual formulae. For now only if there's one donor, because
			I haven't been given a cheatsheet for 2+ donors.
		'''
		percent_host_col = 2*caa - 1		# col num where formula goes. Actual Excel formula, not the text version.
		d1_col = percent_host_col - 4		# col num for donor allele 1
		d2_col = percent_host_col - 3		# col num for donor allele 2
		h1_col = percent_host_col - 2		# col num for host allele 1
		h2_col = percent_host_col - 1		# col num for host allele 2
		for r2 in range(4, 4+2*len(markers), 2):
			r1 = r2 - 1

			d1_alle = ws.cell(row=r1, column=d1_col)
			d2_alle = ws.cell(row=r1, column=d2_col)
			h1_alle = ws.cell(row=r1, column=h1_col)
			h2_alle = ws.cell(row=r1, column=h2_col)

			d1_alle_val = format_value(d1_alle.value)
			d2_alle_val = format_value(d2_alle.value)
			h1_alle_val = format_value(h1_alle.value)
			h2_alle_val = format_value(h2_alle.value)

			d1_area = ws.cell(row=r2, column=d1_col)
			d2_area = ws.cell(row=r2, column=d2_col)
			h1_area = ws.cell(row=r2, column=h1_col)
			h2_area = ws.cell(row=r2, column=h2_col)

			f1 = ws.cell(row=r1, column=percent_host_col+1)
			f1.alignment = Alignment(horizontal='center')
			f2 = ws.cell(row=r2, column=percent_host_col+1)
			f2.alignment = Alignment(horizontal='left')

			d = {d1_alle_val, d2_alle_val}
			d.discard(None)
			h = {h1_alle_val, h2_alle_val}
			h.discard(None)

			percent_host = ws.cell(row=r2, column=percent_host_col)

			# print('d = {}, h = {}'.format(d,h))

			if len(d | h) == 4:
				f1.value = '{} + {}'.format(h1_alle_val,
											h2_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {} + {} + {}'.format(h1_alle_val,
														h2_alle_val,
														d1_alle_val,
														d2_alle_val)

				percent_host.value = '=100*SUM({}:{})/SUM({},{},{},{})'.format(h1_area.coordinate,
																	h2_area.coordinate,
																	d1_area.coordinate,
																	d2_area.coordinate,
																	h1_area.coordinate,
																	h2_area.coordinate)

			if len(h) == 1 and len(h | d) == 3:
				f1.value = '{}'.format(h1_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {} + {}'.format(h1_alle_val,
												d1_alle_val,
												d2_alle_val)

				percent_host.value = '=100*{}/SUM({},{},{})'.format(h1_area.coordinate,
															d1_area.coordinate,
															d2_area.coordinate,
															h1_area.coordinate)

			if len(d) == 1 and len( h | d) == 3:
				f1.value = '{} + {}'.format(h1_alle_val,
											h2_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {} + {}'.format(h1_alle_val,
												h2_alle_val,
												d1_alle_val)

				percent_host.value = '=100*SUM({}:{})/SUM({},{},{})'.format(h1_area.coordinate,
																	h2_area.coordinate,
																	d1_area.coordinate,
																	h1_area.coordinate,
																	h2_area.coordinate)

			if len(h) == 1 and len(d) == 1 and len(h | d) == 2:
				f1.value = '{}'.format(h1_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {}'.format(h1_alle_val,
											d1_alle_val)
				f2.alignment = Alignment(horizontal='center')

				percent_host = '={}/SUM({},{})'.format(h1_area.coordinate,
													d1_area.coordinate,
													h1_area.coordinate)

			if len(h) == 2 and len(d) == 2 and len(h | d) == 3:
				if h1_alle_val not in d:
					h_unique_alle_val = h1_alle_val
					h_unique_area = h1_area
				else:
					h_unique_alle_val = h2_alle_val
					h_unique_area = h2_area

				if d1_alle_val not in h:
					d_unique_alle_val = d1_alle_val
					d_unique_area = d1_area
				else:
					d_unique_alle_val = d2_alle_val
					d_unique_area = d2_area

				f1.value = '{}'.format(h_unique_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {}'.format(h_unique_alle_val,
											d_unique_alle_val)
				f2.alignment = Alignment(horizontal='center')

				percent_host.value = '=100*{}/SUM({},{})'.format(h_unique_area.coordinate,
													d_unique_area.coordinate,
													h_unique_area.coordinate)

			if len(h) == 1 and len(d) == 2 and len(h | d) == 2:
				if d1_alle_val in h:
					A_alle_val = d2_alle_val
					A_area = d2_area
				else:
					A_alle_val = d1_alle_val
					A_area = d1_area

				if h1_alle_val in d:
					H_alle_val = h1_alle_val
					H_area = h1_area
				else:
					H_alle_val = h2_alle_val
					H_area = h2_area

				f2.value = '1 - (2x{}/({} + {}))'.format(A_alle_val,
														A_alle_val,
														H_alle_val)

				percent_host.value = '=100*(1-(2*{}/({}+{})))'.format(A_area.coordinate,
																	A_area.coordinate,
																	H_area.coordinate)
			
			if len(h) == 2 and len(d) == 1 and len(h | d) == 2:
				if h1_alle_val in d:
					A_alle_val = h2_alle_val
					A_area = h2_area
				else:
					A_alle_val = h1_alle_val
					A_area = h1_area

				if d1_alle_val in h:
					D_alle_val = d1_alle_val
					D_area = d1_area
				else:
					D_alle_val = d2_alle_val
					D_area = d2_area

				f1.value = '2x{}'.format(A_alle_val)
				f1.alignment = Alignment(horizontal='center')

				f2.value = '{} + {}'.format(A_alle_val,
											D_alle_val)
				f2.alignment = Alignment(horizontal='center')

				percent_host.value = '=100*(2*{})/({}+{})'.format(A_area.coordinate,
																	A_area.coordinate,
																	D_area.coordinate)

		'''	Formula for average of Percent Host column
		'''
		percent_host_avg = ws.cell(row=raa + 2*len(markers), column=percent_host_col)
		start = ws.cell(row=raa, column=percent_host_col)
		end = ws.cell(row=raa - 1 + 2*len(markers), column=percent_host_col)
		percent_host_avg.value = '=AVERAGE({}:{})'.format(start.coordinate, end.coordinate)

		percent_donor_avg = ws.cell(row=raa + 1 + 2*len(markers), column=percent_host_col)
		percent_donor_avg.value = '=100-{}'.format(percent_host_avg.coordinate)

	return wb

def make_results_dict_from_template(template):
	# print('Now inside makeshift_results_dictionary')
	results_dict = {}
	host_case = None
	# donor_case = None
	donor_cases = []
	host_cell = None
	# donor_cell = None
	donor_cells = []

	if isinstance(template, pd.DataFrame):
		wb = df_to_wb(template)
	elif isinstance(template, openpyxl.Workbook):
		pass
	elif os.path.isfile(str(template)) and template.endswith('.xlsx'):
		# print(template_path)
		wb = openpyxl.load_workbook(template)
		ws = wb.worksheets[0]
	elif os.path.isfile(str(template)) and template.endswith('.xls'):
		template = convert_xls_to_xlsx(template)
		wb = openpyxl.load_workbook(template)

	ws = wb.worksheets[0]

	replacement_dict_no_regex = {
		'THO1': 'TH01',
		'Recipient (Host) Alleles': 'Host',
	}
	replacement_dict_yes_regex = {
		'[vV][wW][aA]':'vWA',
		'[aA][mM][eE][lL][oO][gG][eE][nN][iI][nN]':'AMEL',
	}

	ws = replace_cell_values(ws, replacement_dict_no_regex, regex=False)
	ws = replace_cell_values(ws, replacement_dict_yes_regex, regex=True)

	'''	Get loc of Allele '''
	allele_cell = first_cell_with_value(ws, 'Allele')
	caa = allele_cell.column
	raa = allele_cell.row

	'''	First get all cells that have the word Host/Donor in them '''
	host_cells = all_cells_with_value(ws, r'Host', regex=True)
	donor_cells = all_cells_with_value(ws, r'Donor', regex=True)

	''' Now only keep the host/donor cells that are before and above '''
	host_cells = [cell for cell in host_cells if cell.row < raa and cell.column < caa]
	donor_cells = [cell for cell in donor_cells if cell.row < raa and cell.column < caa]

	'''	Make sure there's only one qualifying host cell '''
	# for cell in host_cells:
	# 	print('host cell coordinate = {}'.format(cell.coordinate))
	assert len(host_cells) == 1, 'Messy template. More than one Host column found between Markers column and Allele/Area column'
	host_cell = host_cells[0]

	'''	Check if it's old style or new style. Case number's location depends on this '''
	old = 'unknown'
	if raa - host_cell.row == 1:
		old = False
		host_case = str(ws.cell(row=host_cell.row - 1, column=host_cell.column).value)
		donor_cases = [str(ws.cell(row=cell.row - 1, column=cell.column).value) for cell in donor_cells]
	elif raa - host_cell.row == 2:
		old = True
		host_case = str(ws.cell(row=host_cell.row + 1, column=host_cell.column).value)
		donor_cases = [str(ws.cell(row=cell.row + 1, column=cell.column).value) for cell in donor_cells]
	else:
		print('Problem! Expecting Host and Donor column titles to be either one row or two rows above the Allele/Area cell.')
	# df = pd.DataFrame(ws.values)
	# print(df)

	markers = ['D8S1179', 'D21S11', 'D7S820', 'CSF1PO', 'D3S1358', 'TH01', 'D13S317', 'D16S539', 'D2S1338', 'D19S433', 'vWA', 'TPOX', 'D18S51', 'AMEL', 'D5S818', 'FGA']
	marker_cells = [first_cell_with_value(ws, marker) for marker in markers if first_cell_with_value(ws, marker) is not None]

	for marker_cell in marker_cells:
		marker = str(marker_cell.value)
		# print('now looping through marker cells')
		# print(marker, marker_cell.value)
		host_allele_0 = ws.cell(row=marker_cell.row, column=host_cell.column).value
		host_allele_1 = ws.cell(row=marker_cell.row, column=host_cell.column + 1).value

		host_key = ('Host', host_case)

		if host_allele_0 is not None:
			results_dict[(host_key, marker, host_allele_0)] = None
		if host_allele_1 is not None:
			results_dict[(host_key, marker, host_allele_1)] = None

		for i, (donor_cell, donor_case) in enumerate(zip(donor_cells, donor_cases), start=1):
			donor_allele_0 = ws.cell(row=marker_cell.row, column=donor_cell.column).value
			donor_allele_1 = ws.cell(row=marker_cell.row, column=donor_cell.column + 1).value

			donor_n = '_'.join(['Donor', str(i)])
			donor_key = (donor_n, donor_case)
			if donor_allele_0 is not None:
				results_dict[(donor_key, marker, donor_allele_0)] = None
			if donor_allele_1 is not None:
				results_dict[(donor_key, marker, donor_allele_1)] = None

	wb.close()
	# print('Now looping through results_dict dictionary')
	# for k,v in results_dict.items():
	# 	print(k,v)
	# print('host_case = {}'.format(host_case))
	# print('donor_cases = {}'.format(donor_cases))
	return results_dict, host_case, donor_cases


def load_workbook_range(range_string, ws):
	col_start, col_end = re.findall("[A-Z]+", range_string)

	data_rows = []
	for row in ws[range_string]:
		data_rows.append([cell.value for cell in row])

	return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


def copy_paste_bottom_few_rows(ws_old, ws_new):
	'''	First, we'll drop any empty columns in ws_old '''

	old_area = all_cells_with_value(ws_old, 'Area', regex=False)[-1]

	range_start = 'A' + str(old_area.row + 1)
	range_end = get_column_letter(old_area.column) + str(ws_old.max_row)
	range_string = range_start + ':' + range_end

	df = load_workbook_range(range_string, ws_old)
	col_to_drop = get_col_to_drop(df)
	df.drop(axis=1, columns=col_to_drop, inplace=True)

	new_area = all_cells_with_value(ws_new, 'Area', regex=False)[-1]
	r_start = new_area.row + 1
	c_start = 1

	for i, _ in enumerate(df.index.tolist()):
		for j, _ in enumerate(df.columns.tolist()):
			r = r_start + i
			c = c_start + j
			ws_new.cell(row=r, column=c).value = str(df.iloc[i,j])
	
	return ws_new



def load_template_and_fix(template):
	''' *********************************************
		THIS IS THE OTHER PLACE TO EDIT FOR 2+ DONORS
		*********************************************
		[x]	make_results_dict_from_template needs to return donor_cases (plural)
		[ ]	load_template_and_fix needs to be compatible with multiple donor_cases
	'''

	results_dict, host_case, donor_cases = make_results_dict_from_template(template)

	markers = ['D8S1179', 'D21S11', 'D7S820', 'CSF1PO', 'D3S1358', 'TH01', 'D13S317', 'D16S539', 'D2S1338', 'D19S433', 'vWA', 'TPOX', 'D18S51', 'AMEL', 'D5S818', 'FGA']

	wb = openpyxl.Workbook()
	ws = wb.active

	ws.cell(row=2, column=1).value = 'Marker'
	ws.cell(row=2, column=2).value = 'Host'

	host_case_abbrev = re.sub(r'_PTE.*$', '', str(host_case))
	ws.cell(row=1, column=2).value = host_case_abbrev

	for i, donor_case in enumerate(donor_cases):
		donor_n = '_'.join(['Donor', str(i+1)])
		donor_case_abbrev = re.sub(r'_PTE.*$', '', str(donor_case))
		ws.cell(row=2, column=4+i*2).value = donor_n
		ws.cell(row=1, column=4+i*2).value = donor_case_abbrev


	'''	Markers & Alleles
	'''
	for i, marker in enumerate(markers):

		r = 1+(i+1)*2
		ws.cell(row=r, column=1).value = marker

		alleles = [k[2] for k in results_dict.keys() if 'Host' == k[0][0] and k[1] == marker]

		for j, allele in enumerate(alleles):
			allele = string_to_number(allele)
			c = j+2
			ws.cell(row=r, column=c).value = allele
			print('Host', marker, j, allele, c, r)

		for d, donor_case in enumerate(donor_cases):
			donor_n = '_'.join(['Donor', str(d+1)])
			alleles = [k[2] for k in results_dict.keys() if k[0][0] == donor_n and k[1] == marker]

			for j, allele in enumerate(alleles):
				allele = string_to_number(allele)
				c = 4 + d*2 + j
				ws.cell(row=r, column=c).value = allele
				print(donor_n, marker, j, allele, c, r)

	''' Add in column of Allele/Area
	'''
	caa = 4 + 2*len(donor_cases)
	for i, marker in enumerate(markers):
		r1 = 1+(i+1)*2
		r2 = r1 + 1
		ws.cell(row=r1, column=caa).value = 'Allele'
		ws.cell(row=r2, column=caa).value = 'Area'

	''' Copy columns
	'''
	for i in range(2,caa,2):
		for r in range(2,ws.max_row + 1):
			c1 = 2*caa - (i+1)
			c2 = c1 + 1
			ws.cell(row=r, column=c1, value=ws.cell(row=r, column=i).value)
			ws.cell(row=r, column=c2, value=ws.cell(row=r, column=i+1).value)

	'''	Add in % Host & Forumula columns
	'''
	ws.cell(row=2, column=2*caa - 1, value='% Host')
	ws.cell(row=2, column=2*caa, value='Formula')
	ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Host')
	ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Donor')

	'''	Add the actual formulae. For now only if there's one donor, because
		I haven't been given a cheatsheet for 2+ donors.
	'''
	wb = insert_equations(wb=wb, host_case=host_case, donor_cases=donor_cases)

	ws = wb.active

	# df = pd.DataFrame(ws.values)
	# print(df)

	'''	Insert "Post-T"'''
	allele_cell = first_cell_with_value(ws, 'Allele', regex=False)
	ws.cell(row=allele_cell.row - 2, column=allele_cell.column + 1).value = 'Post-T:'

	wb_old = openpyxl.load_workbook(template, read_only=True)
	ws_old = wb_old.worksheets[0]
	ws = copy_paste_bottom_few_rows(ws_old, ws)

	# df = pd.DataFrame(ws.values)
	# print(df)

	wb.close()

	return wb


def fill_template_with_areas(template=None, sample_name='', res={}):
	# print('Now running build_profile')
	# print(template)
	# pprint(res)
	if isinstance(template, pd.DataFrame):
		df = template.copy(deep=True)
	elif isinstance(template, openpyxl.Workbook):
		ws = template.worksheets[0]
		df = pd.DataFrame(ws.values)
	elif os.path.isfile(str(template)):
		if template.endswith('.xls'):
			template = convert_xls_to_xlsx(template)
		if template.endswith('.xlsx'):
			# print(template_path)
			wb = openpyxl.load_workbook(template)
			ws = wb.worksheets[0]
			
			df = pd.DataFrame(ws.values)
			wb.close()

			wb = openpyxl.load_workbook(template)
			ws = wb.worksheets[0]

			df = pd.DataFrame(ws.values)
	else:
		return pd.DataFrame()	# return an empty dataframe

	replacement_dict_no_regex = {
		'THO1': 'TH01',
		'Recipient (Host) Alleles': 'Host',
	}
	replacement_dict_yes_regex = {
		'[vV][wW][aA]':'vWA',
		'[aA][mM][eE][lL][oO][gG][eE][nN][iI][nN]':'AMEL',
	}
	# df.replace(to_replace='[vV][wW][aA]', value='vWA', regex=True, inplace=True)
	# df.replace(to_replace='[aA][mM][eE][lL][oO][gG][lL][oO][bB][iI][nN]', value='AMEL', regex=True, inplace=True)
	df.replace(to_replace=replacement_dict_no_regex, inplace=True, regex=False)
	df.replace(to_replace=replacement_dict_yes_regex, inplace=True, regex=True)
	# print(df)
	''' Get locations of 'Allele' '''
	allele_ij = []
	for i in df.index:
		for j, v in enumerate(df.iloc[i]):
			if v == 'Allele':
				allele_ij.append([i, j])
	# print(allele_ij)
	# print(df)
	'''	replace cells that reference other cells with the ref cell's value '''
	df = replace_cell_ref_with_value(df)

	'''	fix the formatting of cells that have a trailing zero '''
	for i, j in allele_ij:
		locus = str(df.iloc[i, 0])
		for k in range(1, j):
			x = str(df.iloc[i, j + k])
			if x.endswith('.0'):
				df.iloc[i, j + k] = x.replace('.0','')

	''' Insert area values '''
	# pprint(res)
	for i, j in allele_ij:
		locus = str(df.iloc[i, 0])
		for k in range(1, j):
			x = str(df.iloc[i, j + k])
			key = (str(sample_name), locus, x)
			val = res.get(key, pd.np.nan)
			# print(key, val)
			df.iat[i + 1, j + k] = res.get(key, pd.np.nan)
			# print('\tcoor = {}, value = {}'.format(coor, x))

	'''	Drop empty columns. Note that this should only be done after formulae are replaced with values '''
	col_to_drop = get_col_to_drop(df)
	df.drop(axis=1, columns=col_to_drop, inplace=True)

	'''	Get rid of the remaining 'Unnamed: #' column labels '''
	df.rename(columns=lambda x: re.sub(r'Unnamed.*', '', str(x)), inplace=True)

	''' Add in case_name '''
	wb = df_to_wb(df)
	ws = wb.worksheets[0]
	sample_case_abbrev = re.sub(r'_PTE.*$', '', str(sample_name))
	post_T_cell = first_cell_with_value(ws, 'Post-T:')
	if post_T_cell is not None:
		r = post_T_cell.row
		c = post_T_cell.column
		ws.cell(row=r, column=c+1).value = sample_case_abbrev
	else:
		allele_cell = first_cell_with_value(ws, 'Allele')
		if allele_cell is not None:
			r = allele_cell.row
			c = allele_cell.column
			ws.cell(row=r-2, column=c+1).value = 'Post-T:'
			ws.cell(row=r-2, column=c+2).value = sample_case_abbrev
	df = pd.DataFrame(ws.values)
	# print(df)
	return df


# def main():
# 	owd = os.getcwd()  # original working directory
# 	os.chdir(r'X:\Hospital\Genetics Lab\DNA_Lab\Ghani')
# 	results = build_results_dict()
# 	f = [k[0] for k in results.keys()]
# 	f = set(f)
# 	for x in f:
# 		print(x)

# 	while True:
# 		build_profile_1(results)


# if __name__ == '__main__':
# 	main()