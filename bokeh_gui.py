from bokeh.io import curdoc
from bokeh.layouts import column, row, gridplot
from bokeh.models.widgets import FileInput, DataTable, DateFormatter, TableColumn, RadioButtonGroup, RadioGroup, Select, TextAreaInput, MultiSelect, Div
from bokeh.models import TextInput, Button, ColumnDataSource, CDSView, IndexFilter, Panel, Tabs
from bokeh.server.server import Server

from convert_fsa_to_csv import convert_file, convert_file_content
from fsa import *
from extract_from_genemapper import *

from os.path import basename
import re
import pandas as pd
import copy
import openpyxl

import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory, asksaveasfilename

from openpyxl.styles import Alignment


def bkapp(curdoc):
	pd.set_option('display.max_columns', 20)
	pd.set_option('display.width', 1000)
	pd.set_option('display.max_rows', 50)

	global_dict = {
		'results_files': [],
		'template_path': None,
		'p0_template_path': None,
		'p1_template_path': None,
		'p1_template_df': pd.DataFrame(),
		'p0_host': None,
		'p0_donors': [],
		'p0_current_case': None,
		'p1_current_case': None,
		'p1_host': None,
		'p1_donors':[],
	}

	def reduce_rows(df):
		# print(df)
		df = df.dropna(axis=1, how='all')
		cols = ['Sample File Name', 'Marker','Allele', 'Area']
		# print('********************************************')
		# print(df)
		df = df.dropna(axis=0, how='any', subset=cols)
		df = df.sort_values(by=cols, ascending=True)
		type_dict = {'Area':'int',
					'Size':'float',
					'Height':'int',
					'Data Point':'int'}
		for key, val in type_dict.items():
			if key in df.columns.tolist():
				df = df.astype({key:val})
		return df


	def p0c0_on_donor_change(attrname, old, new):
		# global p0c1_allele_table

		newest = [c for c in new if c not in old]
		if len(newest) > 0:
			global_dict['p0_current_case'] = newest[0]

			p0c1_table_title.text = str(global_dict['p0_current_case'])
			p0c1_allele_table.source.data = source_cases[newest[0]].data
			p0c1_allele_table.source.selected.indices = source_cases[newest[0]].selected.indices[:]
		refresh_p0_template_preview_table()


	def refresh_p0_template_preview_table():
			''' Preview Template '''
			wb = make_template_wb()
			ws = wb.worksheets[0]
			df = pd.DataFrame(ws.values)
			df.loc[-1] = ''
			# print('Inside of refresh_p0_template_preview_table')
			# print(df)
			# df.loc[-1] = df.columns.tolist()
			df.index = df.index + 1
			df.sort_index(inplace=True)
			col_letters = [openpyxl.utils.get_column_letter(int(i)+1) for i in df.columns.tolist()]
			df.columns = col_letters
			df = df.fillna('')
			df_col = df.columns.tolist()
			columns = [TableColumn(field=col, title=col, width=75) for col in df_col[0]]
			columns.extend([TableColumn(field=col, title=col, width=50) for col in df_col[1:-2]])
			columns.extend([TableColumn(field=col, title=col, width=250) for col in df_col[-2:]])
			p0c2_template_table.columns = columns
			p0c2_template_table.source.data = ColumnDataSource(df).data


	def p0c0_on_host_click(attrname, old, new):
		# global source_cases, p0c1_allele_table

		global_dict['p0_current_case'] = new
		p0c1_table_title.text = str(global_dict['p0_current_case'])
		p0c1_allele_table.source.data = source_cases[new].data
		p0c1_allele_table.source.selected.indices = source_cases[new].selected.indices[:]


	def pNc0_on_results_click():
		# global source_cases, df, p0c1_allele_table, df_cases

		root = tk.Tk()
		root.attributes("-topmost", True)
		root.withdraw()		# hide the root tk window
		file_path = askopenfilename(filetypes = (('Text file', '*.txt'),
												('Comma Separated Values','*.csv'),
												('Tab Separated Values','*.tsv'),
											),
									title = 'Choose GeneMapper results file.',
									initialdir=r'X:\Hospital\Genetics Lab\DNA_Lab\3-Oncology Tests\Engraftment\ABI PTE Runs'
								)
		root.destroy()
		if basename(file_path) is not None and basename(file_path) != '' and basename(file_path) not in global_dict['results_files']:
			global_dict['results_files'].append(basename(file_path))
			pNc0_results_text.value = '\n'.join(global_dict['results_files'])

			df_new = use_csv_module(file_path)
			if not df_new.empty:
				df_new = reduce_rows(df_new)

				case_names = sorted(list(set(df_new['Sample File Name'].to_list())))

				for case in case_names:
					df_copy = df_new.loc[df_new['Sample File Name'] == case].copy(deep=True)
					df_copy.reset_index(inplace=True, drop=True)
					df_copy['Selected'] = False
					indices = pre_selected_indices(df_copy)
					df_copy.loc[indices,'Selected'] = True
					df_cases[case] = df_copy
					# print(df_copy)
					source_cases[case] = ColumnDataSource(df_copy)
					source_cases[case].selected.indices = indices

				p0c0_select_host_case.options = list(source_cases.keys())
				p0c0_select_donor_cases.options = list(source_cases.keys())
				p1c0_select_samples.options = list(source_cases.keys())

		if len(global_dict['results_files']) == 1:
			p0c0_select_host_case.value = list(source_cases.keys())[0]


	def pre_selected_indices(df):
		indices = []
		marker_area_max = {}
		markers = sorted(list(set(df['Marker'].tolist())))
		for marker in markers:
			df_marker = df.loc[df['Marker'] == marker]
			marker_area_max[marker] = df_marker['Area'].max()
		for i in df.index.tolist():
			marker = df.iloc[i].loc['Marker']
			area = df.iloc[i].loc['Area']
			if area >= 0.6 * marker_area_max[marker]:
				indices.append(i)
		return indices


	def make_template_wb(file_path=None):

		''' Helper function '''
		def string_to_number(s):
			try:
				s = float(s)
			except:
				pass
			return s

		''' Helper function '''
		def format_value(cell_value):
			if cell_value is not None:
				cell_value_formatted = str(cell_value).replace('.0', '')
			else:
				cell_value_formatted = None
			return cell_value_formatted


		markers = ['D8S1179', 'D21S11', 'D7S820', 'CSF1PO', 'D3S1358', 'TH01', 'D13S317', 'D16S539', 'D2S1338', 'D19S433', 'vWA', 'TPOX', 'D18S51', 'AMEL', 'D5S818', 'FGA']
		host_case = p0c0_select_host_case.value
		df_host = df_cases.get(host_case, pd.DataFrame())
		df_host = df_host.loc[df_host['Selected'] == True]
		# print('df_host')
		# print(df_host)

		donor_cases = p0c0_select_donor_cases.value
		donors = {}
		for donor_case in donor_cases:
			df_donor = df_cases[donor_case]
			df_donor = df_donor.loc[df_donor['Selected'] == True]
			donors[donor_case] = df_donor.copy(deep=True)

		wb = openpyxl.Workbook()
		ws = wb.active
		'''	Header, etc.
		'''
		ws.oddHeader.center.text = p0c0_enter_host_name.value
		# ws.oddHeader.center.size = 14
		ws.cell(row=2, column=1, value='Marker')
		ws.cell(row=2, column=2, value='Host')
		host_case_abbrev = re.sub(r'_PTE.*$', '', host_case)
		ws.cell(row=1, column=2, value=host_case_abbrev)
		for i, donor_case in enumerate(donor_cases):
			donor_case_abbrev = re.sub(r'_PTE.*$', '', donor_case)
			c = 4 + 2*i
			ws.cell(row=1, column=c, value=donor_case_abbrev)
			if len(donor_cases) == 1:
				donor_num = 'Donor'
			else:
				donor_num = 'Donor ' + str(i+1)
			ws.cell(row=2, column=c, value=donor_num)

		'''	Markers & Alleles
		'''
		for i, marker in enumerate(markers):
			r = 1+(i+1)*2
			df_marker = df_host.loc[df_host['Marker'] == marker]
			ws.cell(row=r, column=1, value=marker)
			# print(df_marker)
			for j, allele in enumerate(df_marker['Allele']):
				# print(marker, j, allele)
				allele = string_to_number(allele)
				c = j+2
				ws.cell(row=r, column=c, value=allele)
			
			for j, donor in enumerate(donors.keys()):
				df_donor = donors[donor]
				df_marker = df_donor.loc[df_donor['Marker'] == marker]
				for k, allele in enumerate(df_marker['Allele']):
					allele = string_to_number(allele)
					c = (4 + 2*j) + k
					ws.cell(row=r, column=c, value=allele)

		''' Add in column of Allele/Area
		'''
		caa = 2*len(donors.keys()) + 4
		for i, marker in enumerate(markers):
			r1 = 1+(i+1)*2
			r2 = r1 + 1
			ws.cell(row=r1, column=caa, value='Allele')
			ws.cell(row=r2, column=caa, value='Area')

		''' Copy columns
		'''
		for i in range(2,caa,2):
			for r in range(2,ws.max_row + 1):
				c1 = 2*caa - (i+1)
				c2 = c1 + 1
				ws.cell(row=r, column=c1, value=ws.cell(row=r, column=i).value)
				ws.cell(row=r, column=c2, value=ws.cell(row=r, column=i+1).value)

		'''	Add in % Host & Forumula columns
		'''
		ws.cell(row=2, column=2*caa - 1, value='% Host')
		ws.cell(row=2, column=2*caa, value='Formula')
		ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Host')
		ws.cell(row=ws.max_row+1, column=2*caa-2, value='%Donor')


		if len(donor_cases) == 1:

			'''	Add the actual formulae. For now only if there's one donor, because
				I haven't been given a cheatsheet for 2+ donors.
			'''
			percent_host_col = 2*caa - 1		# col num where formula goes. Actual Excel formula, not the text version.
			d1_col = percent_host_col - 4		# col num for donor allele 1
			d2_col = percent_host_col - 3		# col num for donor allele 2
			h1_col = percent_host_col - 2		# col num for host allele 1
			h2_col = percent_host_col - 1		# col num for host allele 2
			for r2 in range(4, 4+2*len(markers), 2):
				r1 = r2 - 1

				d1_alle = ws.cell(row=r1, column=d1_col)
				d2_alle = ws.cell(row=r1, column=d2_col)
				h1_alle = ws.cell(row=r1, column=h1_col)
				h2_alle = ws.cell(row=r1, column=h2_col)

				d1_alle_val = format_value(d1_alle.value)
				d2_alle_val = format_value(d2_alle.value)
				h1_alle_val = format_value(h1_alle.value)
				h2_alle_val = format_value(h2_alle.value)

				d1_area = ws.cell(row=r2, column=d1_col)
				d2_area = ws.cell(row=r2, column=d2_col)
				h1_area = ws.cell(row=r2, column=h1_col)
				h2_area = ws.cell(row=r2, column=h2_col)

				f1 = ws.cell(row=r1, column=percent_host_col+1)
				f1.alignment = Alignment(horizontal='center')
				f2 = ws.cell(row=r2, column=percent_host_col+1)
				f2.alignment = Alignment(horizontal='left')

				d = {d1_alle_val, d2_alle_val}
				d.discard(None)
				h = {h1_alle_val, h2_alle_val}
				h.discard(None)

				percent_host = ws.cell(row=r2, column=percent_host_col)

				# print('d = {}, h = {}'.format(d,h))

				if len(d | h) == 4:
					f1.value = '{} + {}'.format(h1_alle_val,
												h2_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {} + {} + {}'.format(h1_alle_val,
															h2_alle_val,
															d1_alle_val,
															d2_alle_val)

					percent_host.value = '=100*SUM({}:{})/SUM({},{},{},{})'.format(h1_area.coordinate,
																		h2_area.coordinate,
																		d1_area.coordinate,
																		d2_area.coordinate,
																		h1_area.coordinate,
																		h2_area.coordinate)

				if len(h) == 1 and len(h | d) == 3:
					f1.value = '{}'.format(h1_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {} + {}'.format(h1_alle_val,
													d1_alle_val,
													d2_alle_val)

					percent_host.value = '=100*{}/SUM({},{},{})'.format(h1_area.coordinate,
																d1_area.coordinate,
																d2_area.coordinate,
																h1_area.coordinate)

				if len(d) == 1 and len( h | d) == 3:
					f1.value = '{} + {}'.format(h1_alle_val,
												h2_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {} + {}'.format(h1_alle_val,
													h2_alle_val,
													d1_alle_val)

					percent_host.value = '=100*SUM({}:{})/SUM({},{},{})'.format(h1_area.coordinate,
																		h2_area.coordinate,
																		d1_area.coordinate,
																		h1_area.coordinate,
																		h2_area.coordinate)

				if len(h) == 1 and len(d) == 1 and len(h | d) == 2:
					f1.value = '{}'.format(h1_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {}'.format(h1_alle_val,
												d1_alle_val)
					f2.alignment = Alignment(horizontal='center')

					percent_host = '={}/SUM({},{})'.format(h1_area.coordinate,
														d1_area.coordinate,
														h1_area.coordinate)

				if len(h) == 2 and len(d) == 2 and len(h | d) == 3:
					if h1_alle_val not in d:
						h_unique_alle_val = h1_alle_val
						h_unique_area = h1_area
					else:
						h_unique_alle_val = h2_alle_val
						h_unique_area = h2_area

					if d1_alle_val not in h:
						d_unique_alle_val = d1_alle_val
						d_unique_area = d1_area
					else:
						d_unique_alle_val = d2_alle_val
						d_unique_area = d2_area

					f1.value = '{}'.format(h_unique_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {}'.format(h_unique_alle_val,
												d_unique_alle_val)
					f2.alignment = Alignment(horizontal='center')

					percent_host.value = '=100*{}/SUM({},{})'.format(h_unique_area.coordinate,
														d_unique_area.coordinate,
														h_unique_area.coordinate)

				if len(h) == 1 and len(d) == 2 and len(h | d) == 2:
					if d1_alle_val in h:
						A_alle_val = d2_alle_val
						A_area = d2_area
					else:
						A_alle_val = d1_alle_val
						A_area = d1_area

					if h1_alle_val in d:
						H_alle_val = h1_alle_val
						H_area = h1_area
					else:
						H_alle_val = h2_alle_val
						H_area = h2_area

					f2.value = '1 - (2x{}/({} + {}))'.format(A_alle_val,
															A_alle_val,
															H_alle_val)

					percent_host.value = '=100*(1-(2*{}/({}+{})))'.format(A_area.coordinate,
																		A_area.coordinate,
																		H_area.coordinate)
				
				if len(h) == 2 and len(d) == 1 and len(h | d) == 2:
					if h1_alle_val in d:
						A_alle_val = h2_alle_val
						A_area = h2_area
					else:
						A_alle_val = h1_alle_val
						A_area = h1_area

					if d1_alle_val in h:
						D_alle_val = d1_alle_val
						D_area = d1_area
					else:
						D_alle_val = d2_alle_val
						D_area = d2_area

					f1.value = '2x{}'.format(A_alle_val)
					f1.alignment = Alignment(horizontal='center')

					f2.value = '{} + {}'.format(A_alle_val,
												D_alle_val)
					f2.alignment = Alignment(horizontal='center')

					percent_host.value = '=100*(2*{})/({}+{})'.format(A_area.coordinate,
																		A_area.coordinate,
																		D_area.coordinate)

			'''	Formula for average of Percent Host column
			'''
			percent_host_avg = ws.cell(row=3+2*len(markers), column=percent_host_col)
			start = ws.cell(row=3, column=percent_host_col)
			end = ws.cell(row=2+2*len(markers), column=percent_host_col)
			percent_host_avg.value = '=AVERAGE({}:{})'.format(start.coordinate, end.coordinate)

			percent_donor_avg = ws.cell(row=4+2*len(markers), column=percent_host_col)
			percent_donor_avg.value = '=100-{}'.format(percent_host_avg.coordinate)


		if file_path is not None:
			if file_path.endswith('.xlsx'):
				'''	Save the file before running fix_formatting '''
				wb.save(file_path)
				wb.close()

				'''	Fix formatting '''
				fix_formatting(file_path)
				print('Done saving {}'.format(file_path))

		wb.close()

		return wb


	def p0c0_on_export_template_click():
		root = tk.Tk()
		root.attributes("-topmost", True)
		root.withdraw()		# hide the root tk window
		file_path = asksaveasfilename(defaultextension = '.xlsx',
										filetypes=[('Excel', '*.xlsx')],
										title = 'Save Template',
										initialfile=p0c0_enter_host_name.value)
		root.destroy()
		if file_path.endswith('.xlsx'):
			make_template_wb(file_path)


	def p0c1_on_select_alleles_change(attrname, old, new):
		# global df_cases

		indices = p0c1_allele_table.source.selected.indices[:]
		source_cases[global_dict['p0_current_case']].selected.indices = indices[:]
		df_cases[global_dict['p0_current_case']].loc[:,'Selected'] = False
		df_cases[global_dict['p0_current_case']].loc[indices,'Selected'] = True
		refresh_p0_template_preview_table()


	def p1c0_on_select_template_click():
		root = tk.Tk()
		root.attributes("-topmost", True)
		root.withdraw()		# hide the root tk window
		file_path = askopenfilename(filetypes=[
												('Excel', '*.xlsx'),
												('Excel', '*.xls'),
											],
									title = 'Choose Template',
									initialdir=r'X:\Hospital\Genetics Lab\DNA_Lab\3-Oncology Tests\Engraftment\Allele Charts')
		root.destroy()

		# print('file_path = {}'.format(file_path))

		''' Convert xls to xlsx if needed '''
		if file_path.endswith('.xls'):
			file_path = convert_xls_to_xlsx(file_path)
		if file_path is not None and file_path != '':
		# drop_empty_columns_and_adjust_formulae(file_path)
			''' Update template_text box'''
			p1c0_template_text.value = basename(file_path)
			global_dict['template_path'] = file_path
			# global template_path
			# template_path = file_path
			# df = fill_template_with_areas_2(template=global_dict['template_path'])
			# makeshift_results_dictionary(template=global_dict['template_path'])
			wb = make_template_from_existing_template(template=global_dict['template_path'])
			patient_name = get_patient_name_from_header(global_dict['template_path'])
			# print('patient_name = {}'.format(patient_name))
			if patient_name is None or patient_name == '':
				template_path = global_dict['template_path']
				patient_name = basename(template_path)
				patient_name = patient_name.replace('.xlsx', '')
				patient_name = patient_name.replace('.xls', '')
			p1c0_enter_host_name.value = patient_name
			df = pd.DataFrame(wb.worksheets[0].values)
			global_dict['p1_template_df'] = df.copy(deep=True)
			'''	But also fill the template if samples are already selected! '''
			# if global_dict['p1_current_case'] is not None:
				# print('\t\t***** NOW WE GONNA REFRESH THIS TABLE \'CAUSE YOLO')
			refresh_p1_template_preview_table()

	def check_if_multiple_donors():
		pass


	def p1c0_on_export_results_click():
		# global template_path
		template_path = global_dict['template_path']

		# header = get_header(template_path)
		# patient_name = header.center.text
		patient_name = p1c0_enter_host_name.value

		for sample in p1c0_select_samples.value:
			'''	Construct output file name '''
			output_file_name = patient_name + ' ' + re.sub(r'_PTE.*$', '', sample)
			root = tk.Tk()
			root.attributes("-topmost", True)
			root.withdraw()		# hide the root tk window
			output_file_name = asksaveasfilename(defaultextension='.xlsx',
											filetypes=[('Excel', '*.xlsx')],
											title='Populate results for {}'.format(sample),
											initialfile = output_file_name)
			root.destroy()

			if output_file_name is not None and output_file_name != '':
				if output_file_name.endswith('.xlsx'):
					pass
				elif output_file_name.endswith('.xls'):
					output_file_name = output_file_name + 'x'
				else:
					output_file_name = output_file_name + '.xlsx'

				df = df_cases[sample]
				# print(df)
				results = build_results_dict(df)
				# print(results)
				df_filled = fill_template_with_areas(res=results, sample_name=sample, template=global_dict['p1_template_df'])
				# global_dict['p1_template_df'] = df_filled.copy(deep=True)
				# print(df_filled)

				# col_letters = [openpyxl.utils.get_column_letter(int(i)+1) for i in df_filled.columns.tolist()]
				# df_filled.columns = col_letters
				df_filled = df_filled.fillna('')
				df_filled.to_excel(output_file_name, index=False, header=False)
				fix_formatting(filename=output_file_name, patient_name=p1c0_enter_host_name.value)


	def p1c0_on_select_samples_change(attrname, old, new):
		# global p1_current_case

		template_path = global_dict['template_path']

		newest = [c for c in new if c not in old]
		if len(newest) > 0:
			'''	Update allele table '''
			global_dict['p1_current_case'] = newest[0]
			refresh_p1_allele_table()
			refresh_p1_template_preview_table()

	def refresh_p1_allele_table():
		if len(str(global_dict['p1_current_case'])) > 0:
			p1c1_table_title.text = str(global_dict['p1_current_case'])
		else:
			p1c1_table_title.text = '&lt;sample&gt;'
		source = source_cases.get('p1_current_case',ColumnDataSource())
		p1c1_allele_table.source.data = source_cases[global_dict['p1_current_case']].data
		# p1c1_allele_table.source.data = source.data
		p1c1_allele_table.source.selected.indices = source_cases[global_dict['p1_current_case']].selected.indices[:]
		# p1c1_allele_table.source.selected.indices = source.selected.indices[:]


	def refresh_p1_template_preview_table():

		df = df_cases.get(global_dict['p1_current_case'],pd.DataFrame())
		# print("df_cases[global_dict['p1_current_case']]")
		# print(df)
		results = build_results_dict(df)
		# pprint(results)
		df_filled = fill_template_with_areas(res=results,
									sample_name=global_dict['p1_current_case'],
									template=global_dict['p1_template_df'])
		global_dict['p1_template_df'] = df_filled.copy(deep=True)
		# print('df_filled')
		# print(df_filled)
		if not df_filled.empty:
			# print(df_filled)
			df_filled.loc[-1] = ''
			# print(df_filled)
			# print(df_filled.index.tolist())
			df_filled.index = df_filled.index + 1
			df_filled.sort_index(inplace=True)

			col_letters = [openpyxl.utils.get_column_letter(int(i)+1) for i in df_filled.columns.tolist()]
			df_filled.columns = col_letters
			df_filled = df_filled.fillna('')
			df_col = df_filled.columns.tolist()
			columns = [TableColumn(field=col, title=col, width=75) for col in df_col[0]]
			columns.extend([TableColumn(field=col, title=col, width=50) for col in df_col[1:-2]])
			columns.extend([TableColumn(field=col, title=col, width=250) for col in df_col[-2:]])
			p1c2_template_table.columns = columns
			p1c2_template_table.source.data = ColumnDataSource(df_filled).data


	def p1c1_on_allele_table_edit(attrname, old, new):
		# print(p1c1_allele_table.source.data)
		pass


	# results_files = []
	source_cases = {}
	df_cases = {}

	pNc0_select_results = Button(label='Add GeneMapper Results', button_type='success')
	pNc0_select_results.on_click(pNc0_on_results_click)
	pNc0_results_text = TextAreaInput(value='<results file>', disabled=True, rows=5)


	# select_host_case = Select(title='Select Host', options=list(source_cases.keys()))
	p0c0_select_host_case = Select(title='Select Host', options=[])
	p0c0_select_host_case.on_change('value', p0c0_on_host_click)


	p0c0_select_donor_cases = MultiSelect(title='Select Donor(s) <ctrl+click to multiselect>',
									options=[],
									size=20)
	p0c0_select_donor_cases.on_change('value', p0c0_on_donor_change)


	p0c0_export_template = Button(label='Export Template', button_type='warning')
	p0c0_export_template.on_click(p0c0_on_export_template_click)


	p0c0_enter_host_name = TextInput(title='Enter Host Name', value='<type host name here>')
	p1c0_enter_host_name = TextInput(title='Enter Host Name', value='<type host name here>')


	p1c0_select_template = Button(label='Select Template', button_type='success')
	p1c0_select_template.on_click(p1c0_on_select_template_click)
	p1c0_template_text = TextAreaInput(value='<template file>', disabled=True)


	p1c0_select_samples = MultiSelect(title='Select Sample(s) <ctrl+click to multiselect>',
										options=[],
										size=20)
	p1c0_select_samples.on_change('value', p1c0_on_select_samples_change)

	p1c0_export_results = Button(label='Export Results To Excel File', button_type='warning')
	p1c0_export_results.on_click(p1c0_on_export_results_click)

	p1c1_table_title = Div(text='&lt;sample&gt;', sizing_mode='fixed')


	p0c1_table_title = Div(text='&lt;sample&gt;', sizing_mode='fixed')
	p0c2_table_title = Div(text='Template', sizing_mode='fixed')
	p1c2_table_title = Div(text='Template', sizing_mode='fixed')

	# p1c2_table_title = Div(text='&lt;sample&gt;', sizing_mode='fixed')


	columns = [#TableColumn(field='Sample File Name', title='Sample File Name', width=300),
				TableColumn(field='Marker', title='Marker', width=75),
				TableColumn(field='Allele', title='Allele', width=50),
				TableColumn(field='Area', title='Area', width=50)]


	source = ColumnDataSource()
	source.selected.on_change('indices', p0c1_on_select_alleles_change)

	p0c1_allele_table = DataTable(columns=columns, source=source, selectable='checkbox', fit_columns=True, sizing_mode='stretch_height', width=300)
	p0c2_template_table = DataTable(source=ColumnDataSource(), fit_columns=True, sizing_mode='stretch_both')
	p1c2_template_table = DataTable(source=ColumnDataSource(), fit_columns=True, editable=True)

	p1c1_allele_table = DataTable(columns=columns, source=ColumnDataSource(), selectable='checkbox', fit_columns=True, sizing_mode='stretch_height', width=300, editable=True)
	p1c1_allele_table.source.on_change('data', p1c1_on_allele_table_edit)

	p0c0 = column(p0c0_enter_host_name,
					pNc0_select_results,
					pNc0_results_text,
					p0c0_select_host_case,
					p0c0_select_donor_cases,
					p0c0_export_template,
					sizing_mode='fixed')

	p0c1 = column(p0c1_table_title, p0c1_allele_table, sizing_mode='stretch_height')

	p0c2 = column(p0c2_table_title, p0c2_template_table, sizing_mode='stretch_both')
	# p0c2 = column(p0c2_template_table)

	p1c0 = column(p1c0_enter_host_name,
					pNc0_select_results,
					pNc0_results_text,
					p1c0_select_template,
					p1c0_template_text,
					p1c0_select_samples,
					p1c0_export_results)

	p1c1 = column(p1c1_table_title, p1c1_allele_table, sizing_mode='stretch_height')

	p1c2 = column(p1c2_table_title, p1c2_template_table, sizing_mode='scale_height')

	child_0 = row(p0c0, p0c1, p0c2, sizing_mode='stretch_both')
	# child_0 = row(p0c0, p0c1, p0c2)
	child_1 = row(p1c0, p1c1, p1c2, sizing_mode='stretch_height')
	tab1 = Panel(child=child_0, title='Make Template')
	tab2 = Panel(child=child_1, title='Populate Results')
	tabs = Tabs(tabs=[tab1, tab2])
	curdoc.add_root(tabs)
	curdoc.title = 'PTE'


# Setting num_procs here means we can't touch the IOLoop before now, we must
# let Server handle that. If you need to explicitly handle IOLoops then you
# will need to use the lower level BaseServer class.
server = Server({'/': bkapp}, num_procs=1)
server.start()

if __name__ == '__main__':
	print('Opening Bokeh application on http://localhost:5006/')

	server.io_loop.add_callback(server.show, "/")
	server.io_loop.start()